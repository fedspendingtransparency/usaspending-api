import logging
from collections import OrderedDict
from pathlib import Path
from time import perf_counter

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from usaspending_api.common.retrieve_file_from_uri import RetrieveFileFromUri
from usaspending_api.references.models import Rosetta

logger = logging.getLogger("script")


EXCEL_COLUMNS = [
    "Element",
    "Definition",
    "FPDS Data Dictionary Element",
    "Grouping",
    "Domain Values",
    "Domain Values Code Description",
    "Award File",
    "Award Element",
    "Subaward File",
    "Subaward Element",
    "Account File",
    "Account Element",
    "Table",
    "Element",
    "Award File",
    "Award Element",
    "Subaward Element",
]


class Command(BaseCommand):
    help = "Loads an Excel spreadsheet of DATA Act/USAspending data names across the various systems into <>"

    s3_public_url = settings.DATA_DICTIONARY_DOWNLOAD_URL

    def add_arguments(self, parser):
        parser.add_argument("-p", "--path", help="filepath to an Excel spreadsheet to load", default=self.s3_public_url)

    def handle(self, *args, **options):
        script_start_time = perf_counter()
        logger.info("Starting load_rosetta management command")
        logger.info("Loading data from {}".format(options["path"]))

        try:
            local_filepath = Path("temp_local_data_dictionary.xlsx")
            RetrieveFileFromUri(options["path"]).copy(str(local_filepath))
            rosetta_object = extract_data_from_source_file(filepath=local_filepath)

            rosetta_object["metadata"]["download_location"] = options["path"]

            load_xlsx_data_to_model(rosetta_object)

            logger.info("Script completed in {:.2f}s".format(perf_counter() - script_start_time))
        except Exception:
            logger.exception("Exception during file retrieval or parsing")
            # Re-raising an exception because the iniital exception is thrown during this handling, it is logged,
            # but the command continues on to exit gracefully with a exit code of 0 (success).
            # This is causing our django-manage-ad-hoc command to be marked as success even
            # if the Rosetta Dictionary fails to update
            raise SystemExit(1)
        finally:
            if local_filepath.exists():
                local_filepath.unlink()


def extract_data_from_source_file(filepath: str) -> dict:
    file_size = filepath.stat().st_size
    wb = load_workbook(filename=str(filepath))
    sheet = wb["Public"]
    # previously we were using sheet.max_column, however that can sometimes be unreliabled
    # opted for hardcoding the last column to traverse, based on the columns defined in EXCEL_COLUMNS
    last_column = get_column_letter(len(EXCEL_COLUMNS))
    cell_range = "A2:{}2".format(last_column)
    headers = [{"column": cell.column_letter, "value": cell.value} for cell in sheet[cell_range][0]]

    sections = []
    for header in headers:
        section = {"section": sheet[f"{header['column']}1"].value, "colspan": 1}
        if section["section"] is None:
            sections[-1]["colspan"] += 1
        else:
            sections.append(section)

    elements = OrderedDict()
    for i, row in enumerate(sheet.values):
        if i < 2:
            continue
        elements[row[0]] = [r.strip() if (r and r.strip() != "N/A") else None for r in row]

    field_names = list(header["value"] for header in headers)
    row_count = len(elements)
    logger.info("Columns in file: {} with {} rows of elements".format(field_names, row_count))

    return {
        "metadata": {
            "total_rows": row_count,
            "total_columns": len(field_names),
            "total_size": "{:.2f}KB".format(float(file_size) / 1024),
        },
        "headers": headers,
        "sections": sections,
        "data": elements,
    }


@transaction.atomic
def load_xlsx_data_to_model(rosetta_object: dict):
    Rosetta.objects.all().delete()
    json_doc = {
        "metadata": rosetta_object["metadata"],
        "sections": rosetta_object["sections"],
        "headers": [
            {
                "display": header["value"],
                "raw": f"{header['column']}:{header['value'].lower().replace(' ','_')}",
            }
            for header in rosetta_object["headers"]
        ],
        "rows": list(row for row in rosetta_object["data"].values()),
    }

    if [elem["display"] for elem in json_doc["headers"]] != EXCEL_COLUMNS:
        raise Exception(f"Columns in excel file don't match reference in code!")

    rosetta = Rosetta(document_name="api_response", document=json_doc)
    rosetta.save()
