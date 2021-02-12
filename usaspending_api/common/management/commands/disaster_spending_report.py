import csv
import glob
import itertools
import json
import logging
import numpy as np
import os
import pandas as pd
import shutil
import zipfile

from datetime import datetime
from django.core.management.base import BaseCommand
from django.db import connection
from openpyxl import load_workbook
from time import perf_counter

from usaspending_api.common.helpers.fiscal_year_helpers import generate_fiscal_year_and_quarter
from usaspending_api.common.helpers.generic_helper import read_text_file

TAS_XLSX_FILE = "usaspending_api/data/DEFC ABC Pd 6 FY18.xlsx"
ASSISTANCE_SQL = read_text_file("usaspending_api/common/management/sql/disaster_spending_assistance.sql")
CONTRACT_SQL = read_text_file("usaspending_api/common/management/sql/disaster_spending_contracts.sql")

logger = logging.getLogger("script")


def dump_to_csv(filepath, data_lines):
    if len(data_lines) == 0:
        return
    if os.path.exists(filepath):
        os.unlink(filepath)

    with open(filepath, "w") as ofile:
        writer = csv.writer(ofile)
        for row in data_lines:
            writer.writerow(row)


def generate_tas_rendering_label(tas_dict):
    """
    Copied from usaspending_api/accounts/models.py and change to use a dict instead of separate parameters
    """
    tas_rendering_label = "-".join(filter(None, (tas_dict["ATA"], tas_dict["AID"])))

    if tas_dict["AvailType Code"] is not None and tas_dict["AvailType Code"] != "":
        tas_rendering_label = "-".join(filter(None, (tas_rendering_label, tas_dict["AvailType Code"])))
    else:
        poa = "\\".join(filter(None, (tas_dict["BPOA"], tas_dict["EPOA"])))
        tas_rendering_label = "-".join(filter(None, (tas_rendering_label, poa)))

    tas_rendering_label = "-".join(filter(None, (tas_rendering_label, tas_dict["Main"], tas_dict["Sub"])))

    return tas_rendering_label


class Command(BaseCommand):
    """
    Management Command to generate a "more expansive" Custom Account Download for specific TAS
    Essentially, it is Account + Award data for each TAS
      - It needs a TAS list (initially provided as an XLSX and stored in usaspending_api/data/)
      - Will create two files for each fiscal quarter of data (one for contracts, another for assistance)
      - Zips the CSVs and compresses for easier transport
    """

    help = "Generate CSV files for provided TAS to help track disaster spending"

    def add_arguments(self, parser):
        default_dir = os.path.dirname(os.path.abspath(__file__))
        parser.add_argument("-d", "--destination", default=default_dir, type=str, help="Location of output file")
        parser.add_argument("-k", "--keep-files", action="store_true", help="If provided, don't delete the temp files")
        parser.add_argument("-p", "--print-zip-path", action="store_true", help="Return the zip path and exit")

    def handle(self, *args, **options):
        script_start = perf_counter()
        self.contract_columns = []
        self.assistance_columns = []
        self.temporary_dir = os.path.dirname(options["destination"] + "/temp_disaster_tas_csv/")
        self.destination = os.path.dirname(options["destination"] + "/OMB_DHS_disaster_report/")
        self.keep_files = options["keep_files"]
        self.verbose = True if options["verbosity"] > 1 else False
        self.zip_filepath = "{}_{}.zip".format(self.destination, datetime.utcnow().strftime("%Y%m%d%H%M_utc"))

        if options["print_zip_path"]:
            print(self.zip_filepath)
            raise SystemExit

        logger.info("Starting Disaster Spending Report script...")

        if not os.path.exists(self.temporary_dir):
            os.makedirs(self.temporary_dir)
        if not os.path.exists(self.destination):
            os.makedirs(self.destination)

        tas_dict_list = self.gather_tas_from_file()
        self.query_database(tas_dict_list)
        self.assemble_csv_files()
        self.cleanup_files()

        logger.info("Success! New zip file: {}".format(self.zip_filepath))
        logger.info("Script completed in {}s".format(perf_counter() - script_start))

    def gather_tas_from_file(self):
        wb = load_workbook(filename=TAS_XLSX_FILE, read_only=True)
        ws = wb["DEFC"]
        ws.calculate_dimension()
        tas_code_header = ws["A8":"G8"]
        expected_headers = ["ATA", "AID", "BPOA", "EPOA", "AvailType Code", "Main", "Sub"]
        headers = []
        for header in tas_code_header:
            for cell in header:
                headers.append(cell.value.replace("\n", "").strip())
        if expected_headers != headers:
            raise Exception("Headers {} Don't match expected: {}".format(headers, expected_headers))

        tas_code_rows = ws["A9":"G100"]
        # Turn "rows" of "cell" into a list of dictionaries using the headers. I'm sure pandas could have done it too
        tas_dicts = [
            {key: cell.value.strip() if cell.value else None for key, cell in itertools.zip_longest(headers, row)}
            for row in tas_code_rows
        ]

        return tas_dicts

    def query_database(self, tas_dict_list):
        db_start = perf_counter()

        for i, tas in enumerate(tas_dict_list, 1):
            contract_results = self.single_tas_query(tas, "contract")
            filepath = "{}/{}_fpds.csv".format(self.temporary_dir, generate_tas_rendering_label(tas))
            dump_to_csv(filepath, contract_results)

            assistance_results = self.single_tas_query(tas, "assistance")
            filepath = "{}/{}_fabs.csv".format(self.temporary_dir, generate_tas_rendering_label(tas))
            dump_to_csv(filepath, assistance_results)

            if self.verbose and (len(tas_dict_list) - i) % 10 == 0:
                logger.info("{:2>} TAS left".format(len(tas_dict_list) - i))

        if self.verbose:
            logger.info(self.contract_columns)
            logger.info(self.assistance_columns)

        logger.info("Completed all database queries in {}s".format(perf_counter() - db_start))

    def single_tas_query(self, tas_dict, transaction_type="contract"):
        single_db_query = perf_counter()
        if transaction_type == "contract":
            sql_string = CONTRACT_SQL
        else:
            sql_string = ASSISTANCE_SQL
        formatted_dict = {k: "= '{}'".format(v) if v else "IS NULL" for k, v in tas_dict.items()}
        sql_string = sql_string.format(**formatted_dict)

        results = []
        with connection.cursor() as cursor:
            cursor.execute(sql_string)
            if transaction_type == "contract" and not self.contract_columns:
                self.contract_columns = [col[0] for col in cursor.description]
            elif transaction_type == "assistance" and not self.assistance_columns:
                self.assistance_columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
        if self.verbose:
            logger.info(json.dumps(tas_dict))
            logger.info(
                "Query for {}s using above TAS took {}s returning {} rows".format(
                    transaction_type, perf_counter() - single_db_query, len(results)
                )
            )
        return results

    def assemble_csv_files(self):
        logger.info("Using pandas to read temporary .csv files")
        start_pandas = perf_counter()

        files = glob.glob(self.temporary_dir + "/*_fpds.csv")
        df = pd.concat([pd.read_csv(f, dtype=str, header=None, names=self.contract_columns) for f in files])
        self.write_pandas(df, "fpds")

        files = glob.glob(self.temporary_dir + "/*_fabs.csv")
        df = pd.concat([pd.read_csv(f, dtype=str, header=None, names=self.assistance_columns) for f in files])
        self.write_pandas(df, "fabs")

        logger.info("Assembling data and saving to final .csv files took {}s".format(perf_counter() - start_pandas))

    def write_pandas(self, df, award_type):
        df = df.replace({np.nan: None})
        df = df.replace({"NaN": None})
        df["submission_period"] = pd.to_datetime(df["submission_period"])
        df["_fyq"] = df.apply(lambda x: generate_fiscal_year_and_quarter(x["submission_period"]), axis=1)

        if self.verbose:
            logger.info("Completed pandas dataframe for all {} records".format(award_type))

        for quarter in pd.unique(df["_fyq"]):
            filepath = "{}/{}_{}.csv".format(self.destination, quarter, award_type)
            temp_df = df.loc[df["_fyq"] == quarter]
            del temp_df["_fyq"]
            temp_df.to_csv(path_or_buf=filepath, index=False)

    def cleanup_files(self):
        csv_files = glob.glob(self.destination + "/*.csv")
        zipped_csvs = zipfile.ZipFile(self.zip_filepath, "a", compression=zipfile.ZIP_DEFLATED, allowZip64=True)

        for csv_file in csv_files:
            zipped_csvs.write(csv_file, os.path.basename(csv_file))

        if not self.keep_files:
            logger.info("Removing temporary directories along with temporary files")
            shutil.rmtree(self.temporary_dir)
            shutil.rmtree(self.destination)
